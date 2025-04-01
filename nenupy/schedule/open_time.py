#! /usr/bin/python3
# -*- coding: utf-8 -*-


"""
    *******************
    Open time functions
    *******************

    TODO
    - check that there is no overlap in a single KP's input
    - BUG wrong display event width when overlapping several other non-overlapping events
"""


__author__ = "Alan Loh"
__copyright__ = "Copyright 2024, nenupy"
__credits__ = ["Alan Loh"]
__maintainer__ = "Alan"
__email__ = "alan.loh@obspm.fr"
__status__ = "Production"
__all__ = [
    "XLSFile",
    "NenuCalendar"
]



import numpy as np
import re
import functools
from itertools import tee, islice, chain
import operator
import os
from typing import List, Dict, Tuple
from ics import Calendar, Event
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell, MergedCell
from astropy.time import Time, TimeDelta
import calendar
import astropy.units as u
from astropy.coordinates import Angle
import datetime
import pytz

import matplotlib.dates as mdates
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as bkpdf

import logging
log = logging.getLogger(__name__)
log.setLevel(logging.INFO)


SEASON = {
    "winter_start": "09-30",
    "summer_start": "03-31"
}
NIGHT_TIME = {
    "winter": (Angle([0, 4], unit="hourangle"), Angle([20, 24], unit="hourangle")), # local hours
    "summer": (Angle([0, 3], unit="hourangle"), Angle([19, 24], unit="hourangle"))
}


def _season_from_time(day: Time) -> str:
    year = day.yday.split(":")[0]

    winter_start = Time(f"{year}-{SEASON['winter_start']}")
    summer_start = Time(f"{year}-{SEASON['summer_start']}")

    if (day >= summer_start) and (day < winter_start):
        # We are in 'summer'
        season = "summer"
    else:
        # We are in 'winter'
        season = "winter"
    log.debug(f"{day.isot} is in <{season}> time.")
    return season

def _sort_night_time(start_hour: Angle, stop_hour: Angle, current_day: Time) -> Tuple[float, float]:
    
    season = _season_from_time(current_day)

    morning_night_boundaries, evening_night_boundaries = NIGHT_TIME[season]

    day_time = (stop_hour - start_hour).hour
    morning_hours = np.max((0, (morning_night_boundaries[1] - start_hour).hour))
    evening_hours = np.max((0, (stop_hour - evening_night_boundaries[0]).hour))
    night_time = morning_hours + evening_hours
    night_time += np.min((0, (day_time - night_time))) # if only morning or evening, correct the equation
    day_time -= night_time

    return day_time, night_time

# ============================================================= #
# -------------------------- XLSFile -------------------------- #
class XLSFile:

    def __init__(self, filename: str):
        self.filename = filename
        self.data = {}

        # Load data

        self.workbook = openpyxl.load_workbook(filename)
        self.sheets = self._read_sheet()

        for sheet in self.sheets:
            sheet_name = sheet.title
            dates, month_data = self._parse_month(sheet)
            self.data[sheet_name] = {}
            self.data[sheet_name]["time"] = dates
            self.data[sheet_name]["data"] = month_data
        
        log.info(f"{filename} read and parsed.")

    def to_ical(self) -> Dict[str, Calendar]:
        # Create a calendar per KP
        calendars = {}

        # Gather all data
        days = functools.reduce(operator.add, [self.data[month_key]["time"] for month_key in self.data.keys()])
        data = np.vstack( [self.data[month_key]["data"] for month_key in self.data.keys()] )

        # Fill calendars while looping through data
        for day, daily_data in zip(days, data):
            # Find cells that have been filled
            values = np.unique(daily_data[daily_data != None])
            for value in values:
                fullmatch = re.search(pattern=r"^(?P<kp>(ES|SP|RP|LT)\S{2})\s+(?P<start>(\d+|\d+:\d+))-(?P<stop>(\d+|\d+:\d+))(?P<comment>\s+\(.+\))?(\s+)?$", string=value)
                if fullmatch is None:
                    print(f"Problem parsing: '{value}'")
                # Add a new calendar if a different KP name is found
                if fullmatch["kp"] not in calendars:
                    calendars[fullmatch["kp"]] = Calendar()
                start_time = day + datetime.timedelta(hours=Angle(fullmatch["start"], unit="hourangle").hour)
                stop_time = day + datetime.timedelta(hours=Angle(fullmatch["stop"], unit="hourangle").hour)
                # Add an event to the corresponding calendar
                calendars[fullmatch["kp"]].events.add(
                    self._new_calendar_event(
                        name=value,
                        start=start_time,
                        stop=stop_time,
                        comment=fullmatch["comment"]
                    )
                )

        for kp in calendars:            
            # Group bookings before / after midnight
            # for event in sorted(calendars[kp].events):
            current_events, next_events = tee(sorted(calendars[kp].events), 2)
            next_events = chain(islice(next_events, 1, None), [None])

            for event, next_event in zip(current_events, next_events):
                if next_event is None:
                    # End of the file
                    continue

                comment_search = re.search(pattern=r"\((?P<comment>.+)\)", string=event.name)
                comment = "" if comment_search is None else comment_search["comment"]
                next_comment_search = re.search(pattern=r"\((?P<comment>.+)\)", string=next_event.name)
                next_comment = "" if next_comment_search is None else next_comment_search["comment"]
                if (next_event.begin == event.end) and (next_comment == comment):
                    # Merge the events and remove the old one
                    next_event.begin = event.begin
                    next_event.name = f"{event.name}*{next_event.name}" 
                    calendars[kp].events.remove(event)
    
        return calendars

    def to_ics(self, save_path: str = "") -> None: # Assumes the file is perfect!
        
        calendars = self.to_ical()

        # Write all calendars
        for kp in calendars:
            
            output = os.path.join(save_path, f"{kp}.ics")
    
            with open(output, "w") as wf:
                wf.writelines(calendars[kp].serialize_iter())

    def info(self, kp: str = "", strict: bool = False):
        
        hour_total = 0
        day_total = 0
        night_total = 0
        errors = 0
        n_values = 0

        for month_key in self.data.keys():

            month_days = self.data[month_key]["time"]
            month_data = self.data[month_key]["data"]

            month_hours = 0
            month_night = 0
            month_day = 0

            for day, day_data in zip(month_days, month_data):

                # Get the different cell values and their occurences
                hour_column = np.arange(0, 24)
                values, indices, counts = np.unique(day_data[day_data != None], return_counts=True, return_index=True)
                hour_indices = hour_column[day_data != None][indices]
                n_values += values.size

                for value, index, count in zip(values, hour_indices, counts):

                    # Super strict check
                    try:
                        fullmatch = re.search(pattern=r"^(?P<kp>(ES|SP|RP|LT)\S{2})\s(?P<start>(\d+|\d+:\d+))-(?P<stop>(\d+|\d+:\d+))(?P<comment>\s\(.+\))?$", string=value)
                        assert fullmatch is not None
                    except AssertionError:
                        log.error(f"(Month {month_key}, Day {day}) - Strict syntax check failed on '{value}'")
                        errors += 1
                        if strict:
                            continue
                        else:
                            pass

                    # Read KP
                    try:
                        kp_match = re.match(pattern=r"^(ES|SP|RP|LT)\d+", string=value)
                        # current_kp = value.split(" ")[0]#.replace("(", "").replace(")", "")
                        current_kp = kp_match.group()
                    except AttributeError:
                        log.error(f"(Month {month_key}, Day {day}) - Impossible to read KP '{value}'")
                        errors += 1
                        continue
                    # Skip if not the selected kp
                    if (kp != "") and (current_kp != kp):
                        continue

                    # Read the start and stop hours
                    try:
                        match = re.search(pattern=r"^(ES|SP|RP|LT)\S+ (?P<start>\S+)-(?P<stop>\S+)", string=value)
                        start = match["start"]
                        stop = match["stop"]
                        # start, stop = value.split(" ")[1].split("-")
                    except TypeError:
                        log.error(f"(Month {month_key}, Day {day}) - Impossible to read start/stop times '{value}'")
                        errors += 1
                        continue

                    # Compute delta-time
                    try:
                        start_val = Angle(start, unit="hourangle")
                        stop_val = Angle(stop, unit="hourangle")
                        hours = (stop_val - start_val).hour
                    except:
                        log.error(f"(Month {month_key}, Day {day}) - not understanding start={start} stop={stop} values...")
                        errors += 1
                        continue
                    
                    # Differentiate night from day time
                    day_hours, night_hours = _sort_night_time(
                        start_hour=start_val,
                        stop_hour=stop_val,
                        current_day=Time(day, format="datetime")
                    )

                    # Check that the time duration makes sense
                    # if hours < 0:
                    #     raise ValueError(f"(Month {month_key}, Day {day}) - Negative duration for '{value}'")
                    # elif hours != (day_hours + night_hours):
                    #     raise ValueError(f"(Month {month_key}, Day {day}) - Problem parsing day/night hours...")
                    # elif hours > 24:
                    #     raise ValueError(f"(Month {month_key}, Day {day}) - >24h duration for '{value}'")

                    # Check that the start is at the correct position
                    if int(start_val.hour) != index:
                        errors += 1
                        log.error(f"(Month {month_key}, Day {day}) - Legend hour {start_val.hour} does not match column {index}h")

                    # Check that the time duration corresponds to the number of merged cells
                    if count != int(np.ceil(hours)):
                        errors += 1
                        log.error(
                            f"(Month {month_key}, Day {day}) - {count} cells instead of {int(np.ceil(hours))} for value '{value}'"
                        )
                    
                    # Add the time to the total
                    month_day += day_hours
                    month_night += night_hours
                    month_hours += hours

    
            log.info(f"Month {month_key}: {month_hours} hours (day={month_day}hrs, night={month_night}hrs)")
            hour_total += month_hours
            day_total += month_day
            night_total += month_night

        log.info(f"total={hour_total} hours (day={day_total}, night={night_total})")
        log.info(f"{errors=} / entries={n_values}")

    def _read_sheet(self) -> List[Worksheet]:
        sheets = []
        for sheet_name in self.workbook.sheetnames:
            # Skip the example sheet
            if sheet_name.lower() == "example":
                log.debug(f"'{sheet_name}' sheet skiped.")
                continue
            log.debug(f"Parsing '{sheet_name}' from '{self.filename}'")
            sheets.append(self.workbook[sheet_name])
        if not len(sheets) == 6:
            raise ValueError(f"{len(sheets)} sheets were returned instead of 6 while reading {self.filename}.")
        return sheets

    def _parse_month(self, sheet: Worksheet) -> Tuple[list, np.ndarray]:

        # Loop over the rows of the sheet
        sheet_rows = []

        dates = []
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=24 + 1):

            # Only consider rows belonging to a month days
            current_day = row[0].value
            if not isinstance(current_day, datetime.datetime):
                break
            dates.append(current_day)

            sheet_rows.append(self._row_to_list(row[1:]))

        # Build the entire sheet and return a numpy array
        return dates, np.array(sheet_rows)

    @staticmethod
    def _row_to_list(row: Tuple[Cell]) -> list:
        # Find out if and where there are merged cells
        merged_mask = np.array([isinstance(cell, MergedCell) for cell in row], dtype=bool)
        
        # If there are merged cells, compute the indices of the first cells of merged cells (they are not of type MergedCell)
        merged_indices = merged_mask.nonzero()[0]
        try:
            first_cell_indices = np.array([group[0] for group in np.split(merged_indices, np.where(np.diff(merged_indices) != 1)[0]+1)]) - 1
        except IndexError:
            first_cell_indices = np.array([])

        # Loop over each cell in the row
        row_values = []
        for i, cell in enumerate(row):

            if merged_mask[i]:
                # if the cell is merged, get the corresponding value from the list of first cell values
                previous_cell_index = first_cell_indices[first_cell_indices < i][-1]
                row_values.append( row[previous_cell_index].value )

            else:
                row_values.append(cell.value)

        # Make sure that the 24 hours have been correctly read
        if len(row_values) != 24:
            raise ValueError(f"Row {row}: input different than 24 hours..")

        return row_values

    @staticmethod
    def _new_calendar_event(name: str, start: datetime.datetime, stop: datetime.datetime, comment: str = "") -> Event:
        event = Event()
        event.name = name
        event.begin = start
        event.end = stop
        event.transparent = False
        event.description = comment
        return event


# ============================================================= #
# ------------------------- NenuEvent ------------------------- #
SIDEREAL_DAY_TIMEDELTA = datetime.timedelta(hours=23, minutes=56, seconds=4.0905)
# SIDEREAL_DAY_TIMEDELTA = datetime.timedelta(minutes=30, seconds=4.0905)
class NenuEvent:

    def __init__(self, event: Event, kp_name: str, color: str):
        self.event = event
        self.color = color
        self.kp_name = kp_name
        self.draw_box = None
        self.draw_text = None
        self.selected = False

    def __repr__(self) -> str:
        return f"<NenuEvent({self.event.begin.datetime} hello {self.event.end=} {self.kp_name=})>"
    
    def __str__(self) -> str:
        return f"<NenuEvent({self.event.begin.datetime} -- {self.event.end.datetime} KP='{self.kp_name}')>"
    
    def __lt__(self, other):
        return self.event.begin < other.event.end

    def __contains__(self, other) -> bool:
        """Checks whether self and other are sharing a common time interval.
        i.e. if DT_1 + DT_2 > max( |f_1 - i_2|, |f_2 - i_1| ) -->

        Parameters
        ----------
        other : _type_
            _description_

        Returns
        -------
        bool
            _description_
        """
        dt_1 = self.event.end.datetime - self.event.begin.datetime
        dt_2 = other.event.end.datetime - other.event.begin.datetime
        diff_1 = np.abs( self.event.end.datetime - other.event.begin.datetime )
        diff_2 = np.abs( other.event.end.datetime - self.event.begin.datetime )
        
        return dt_1 + dt_2 > np.max( (diff_1, diff_2) )

    def contains(self, mouse_event) -> bool:
        bl, tl, _, br, _ = self.draw_box.get_xy()
        ex, ey = (mouse_event.xdata, mouse_event.ydata)
        return (ex >= bl[0]) & (ex <= br[0]) & (ey >= bl[1]) & (ey <= tl[1])

    def connect_to_plot_events(self):
        if self.draw_box is None:
            raise Exception("Call draw() method first.")
        self.cid_click = self.draw_box.figure.canvas.mpl_connect(
            "button_press_event", self._on_click)
        self.cid_press = self.draw_box.figure.canvas.mpl_connect(
            "key_press_event", self._on_key_press)
        # self.cid_pick = self.draw_box.figure.canvas.mpl_connect(
        #     "pick_event", self._on_pick)
        
        # self.cidpress = self.draw_box.figure.canvas.mpl_connect(
        #     'button_press_event', self.on_press)
        # self.cidrelease = self.draw_box.figure.canvas.mpl_connect(
        #     'button_release_event', self.on_release)
        # self.cidmotion = self.draw_box.figure.canvas.mpl_connect(
        #     'motion_notify_event', self.on_motion)

    def draw(self, ax: mpl.axes.Axes, fontsize: int = 6, text_rotation: int = 90, text_max_size: int = None):
        intersections_before, intersections_after = (0, 0)#self._n_intersections_event(event) TODO reactivate
        n_intersections = intersections_before + intersections_after + 1
        
        y_width = 1 / n_intersections
        ymin = intersections_before * y_width
        ymax = ymin + y_width

        # Show the block rectangle
        self.draw_box = ax.axvspan(
            self.event.begin,
            self.event.end,
            ymin=ymin,
            ymax=ymax,
            facecolor=self.color,
            edgecolor="black",
            alpha=0.8,#0.5,
            zorder=40,
            # picker=True
        )

        # Show the observation block title
        text = self.event.name
        if text_max_size is None:
            pass
        elif (text.lower() == "filler") or (text.lower() == "filer"):
            text = "LT02 Filler time"
        else:
            text = "\n".join(text[i:i+text_max_size] for i in range(0, len(text), text_max_size))

        x_min, x_max = ax.get_xlim()
        text_pos = (self.event.begin + (self.event.end - self.event.begin)/2)
        text_pos_mdate = mdates.date2num(text_pos.datetime)
        if (x_min <= text_pos_mdate) & (text_pos_mdate < x_max):
            self.draw_text = ax.text(
                x=text_pos.datetime,
                y=ymin + y_width / 2,
                s=text,# + f"\n{intersections_before} -- {intersections_after}",
                horizontalalignment="center",
                verticalalignment="center",
                rotation=text_rotation,
                color="black",
                fontsize=fontsize,
                zorder=50
            )
    
    def _erase_drawing(self):
        current_figure = self.draw_box.figure
        print(self.event.name)
        self.draw_box.remove()
        # current_figure.canvas.draw_idle()


    def _shift_event_one_day_earlier(self) -> None:
        # one_day = datetime.timedelta(days=1)
        self.event.begin -= SIDEREAL_DAY_TIMEDELTA
        self.event.end -= SIDEREAL_DAY_TIMEDELTA
        vertices = self.draw_box.get_xy()
        vertices[np.array([0, 1, 4]), 0] = mdates.date2num(self.event.begin)
        vertices[np.array([2, 3]), 0] = mdates.date2num(self.event.end)
        self.draw_box.set_xy(vertices)
        self.draw_text.set_x(mdates.date2num(self.event.begin + (self.event.end - self.event.begin)/2))

    def _shift_event_one_day_later(self) -> None:
        # one_day = datetime.timedelta(days=1)
        self.event.end += SIDEREAL_DAY_TIMEDELTA
        self.event.begin += SIDEREAL_DAY_TIMEDELTA
        vertices = self.draw_box.get_xy()
        vertices[np.array([0, 1, 4]), 0] = mdates.date2num(self.event.begin)
        vertices[np.array([2, 3]), 0] = mdates.date2num(self.event.end)
        self.draw_box.set_xy(vertices)
        self.draw_text.set_x(mdates.date2num(self.event.begin + (self.event.end - self.event.begin)/2))

    def _on_click(self, event):
        # print(self.draw_box.get_xy())
        # contains, attrd = self.draw_box.contains(event)
        # print(self.event.name, f"contains={contains}", f"{attrd}")
        if event.inaxes is None:
            return

        if not self.contains(event):
            if self.selected:
                # If it was previously selected, unselect
                self.selected = False
                self._change_facecolor(color=self.color)
                print(f"Un_selecting {self.event.name} !")
            return

        self.selected = True
        self._change_facecolor(color="gray")
        print(f"Selected {self.event.name} !")
    
    def _on_key_press(self, event):
        if (not self.selected) or (event.key not in ["up", "down"]):
            return
        
        current_ax = self.draw_box.axes

        if event.key == "up":
            print("go to past")
            self._shift_event_one_day_earlier()
        elif event.key == "down":
            print("go to future")
            self._shift_event_one_day_later()
        
        # self._erase_drawing()
        
        self._update_draw()

    def _change_facecolor(self, color) -> None:
        background = self.draw_box.figure.canvas.copy_from_bbox(self.draw_box.axes.bbox)
        self.draw_box.set(facecolor=color)
        self._update_draw(background)
    
    def _update_draw(self, background, ax: mpl.axes.Axes = None) -> None:
        if ax is None:
            ax = self.draw_box.axes
        figure = self.draw_box.figure
        figure.canvas.draw()
        figure.canvas.restore_region(background)
        figure.canvas.blit(ax.bbox)

        #self.draw_box.figure.canvas.draw()
        # if ax is None:
        #     ax = self.draw_box.axes
        #del self.draw_box
        #del self.draw_text
        # self.draw(ax=ax)
        
        # for ax in self.draw_box.figure.axes:
        #     self.draw_box.figure.canvas.blit(ax.bbox)
        # self.draw_box.figure.canvas.draw()
        # self.draw_box.figure.canvas.flush_events()
        # plt.gcf().canvas.draw_idle()
        # for ax in self.draw_box.figure.axes:
        #     ax.draw_idle()
        # self.draw_box.figure.canvas.draw_idle()

# ============================================================= #
# ----------------------- NenuCalendar ------------------------ #
class NenuCalendar:

    def __init__(self, events: List[NenuEvent]):
        self.events = events
    # def __init__(self, calendars: Dict[str, Calendar]):
    #     self.kp_names = list(calendars.keys())
    #     self.calendars = calendars
    #     self.kp_colors = mpl.cm.get_cmap("Spectral")(np.linspace(0, 1, len(self.kp_names)))

    @classmethod
    def from_ics(cls, *ics_files: str):

        calendars = {}

        for ics_file in ics_files:

            # Search the KP name
            ics_basename = os.path.basename(ics_file)
            match = re.search(pattern=r"(ES|SP|RP|LT)\S{2}", string=ics_basename)
            if match is None:
                raise ValueError(f"ICS file name {ics_basename} does not contain a KP name.")
            kp = match.group()

            # Read the ICS file
            with open(ics_file, "r") as rfile:
                ics_content = rfile.read()

            calendars[kp] = Calendar(ics_content)
        
        events = []
        kp_colors = mpl.colormaps["Spectral"](np.linspace(0, 1, len(calendars.keys())))
        for kp, kp_color in zip(calendars.keys(), kp_colors):
            if len(calendars.keys()) == 1:
                kp_color = "tab:blue"
            for event in calendars[kp].events:
                events.append(NenuEvent(event, kp_name=kp, color=kp_color))

        return cls(events)

    @classmethod
    def from_xls(cls, *xls_files: str):
        
        kp_calendars = {}
        
        for xls_file in xls_files:
            
            # Search the KP name
            xls_basename = os.path.basename(xls_file)
            match = re.search(pattern=r"(ES|SP|RP|LT)\S{2}", string=xls_basename)
            if match is None:
                log.warning(f"ICS file name {xls_basename} does not contain a KP name.")
                kp = "none"
            else:
                kp = match.group()

            xls = XLSFile(filename=xls_file)
        
            kp_calendars = {**kp_calendars, **xls.to_ical()}
        
        # return cls(kp_calendars)
        events = []
        kp_colors = mpl.colormaps["Spectral"](np.linspace(0, 1, len(kp_calendars.keys())))
        for kp, kp_color in zip(kp_calendars.keys(), kp_colors):
            for event in kp_calendars[kp].events:
                events.append(NenuEvent(event, kp_name=kp, color=kp_color))

        return cls(events)

    def info(self) -> None:
        # TBD: How much time is scheduled
        
        total_hours = 0
        # daily_hours = 0
        # night_hours = 0

        for evt in self.events:
            # start = evt.event.begin.datetime
            # start_hour = start.hour + start.minute / 60 + start.second / 3600
            # stop = evt.event.end.datetime
            # stop_hour = stop.hour + stop.minute / 60 + stop.second / 3600
            # if stop_hour < start_hour:
            #     # This is the next day
            #     stop_hour += 24
            # start_val = Angle(start_hour, unit="hourangle")
            # stop_val = Angle(stop_hour, unit="hourangle")
            # hours = (stop_val - start_val).hour

            # d_hours, n_hours = _sort_night_time(
            #     start_hour=start_val,
            #     stop_hour=stop_val,
            #     current_day=Time(evt.event.begin.datetime, format="datetime")
            # )

            # daily_hours += d_hours
            # night_hours += n_hours
            td = evt.event.end.datetime - evt.event.begin.datetime
            total_hours += td.days * 24 + td.seconds / 3600

        # log.info(f"total={total_hours} hours (day={daily_hours}, night={night_hours})")
        log.info(f"total={total_hours} hours")# (day={daily_hours}, night={night_hours})")

    def write_vcr_csv(self, path: str = "", add_error: bool = True) -> None:
        """Write a CSV file from the events loaded in this NenuCalendar instance.
        The CSV format is such that it can be imported to book KP slots in the VCR.

        Parameters
        ----------
        path : str, optional
            The output path to store the booking file, by default ""
        add_error : bool, optional
            Duplicater the last line so that it results in an error (i.e., to check that this is the last remaining error while importing the data to the VCR), by default True
        """

        sorted_events = sorted(self.events)
        current_events, next_events = tee(sorted_events, 2)
        next_events = chain(islice(next_events, 1, None), [None])

        yyyy_mm_start = sorted_events[0].event.begin.datetime.strftime("%Y_%m")
        yyyy_mm_stop = sorted_events[-1].event.end.datetime.strftime("%Y_%m")
        filename = os.path.join(path, f"Booking_{yyyy_mm_start}_{yyyy_mm_stop}.csv")

        log.info(f"Writing {filename}...")

        with open(filename, "w") as wfile:
            for event, next_event in zip(current_events, next_events):
                
                if next_event is None:
                    # End of the loop
                    continue

                # Check that there is no overlap!
                if event in next_event:
                    raise Exception(f"There is an overlap between {event} and {next_event}!")

                start = Time(event.event.begin.datetime, format="datetime")
                start.precision = 0
                stop = Time(event.event.end.datetime, format="datetime")

                # Gather events belonging to the same KP that are consectutive in time
                # This will modify the stop time
                while (event.kp_name == next_event.kp_name) and (event.event.end == next_event.event.begin):
                    # Check that there is no overlap!
                    if event in next_event:
                        raise Exception(f"There is an overlap between {event} and {next_event}!")
                    if next_event is None:
                        # End of the loop
                        break
                    log.info(f"Merging {event} and {next_event}")
                    stop = Time(next_event.event.end.datetime, format="datetime")
                    event = next(current_events)
                    next_event = next(next_events)

                stop.precision = 0

                # TODO Remove the conditions next time...
                event_name = "LT02 filler" if (event.event.name.lower() == "filler" or event.event.name.lower() == "filer") else event.event.name
                if "LT00" in event.event.name:
                    event_name = event_name.replace("LT00", "ES00")
                if "RP5A" in event.event.name:
                    event_name = event_name.replace("RP5A", "RP3A")

                # Extract the KP name
                kp_match = re.search(pattern=r"(ES|SP|RP|LT)\S{2}", string=event_name)
                if kp_match is None:
                    raise Exception(f"Unable to read KP name from '{event_name}'")
                kp = kp_match.group()
                if kp != event.kp_name:
                    raise Exception(f"Inconsistency between kp_name {event.kp_name} and kp written {kp}")

                # Write the CSV file
                wfile.write(f"{start.iso},{stop.iso},{kp},NenuFAR booking\n")

            if add_error:
                # Duplicate last line to generate an error
                wfile.write(f"{start.iso},{stop.iso},{kp},NenuFAR booking\n")

    def month_plot(self, fig_name: str):

        if not fig_name.endswith(".pdf"):
            raise ValueError("fig_name must end with .pdf")

        pdf_document = bkpdf.PdfPages(fig_name)

        utc = pytz.UTC

        # Sort events by month
        monthly_events = {}

        for evt in self.events:
            begin_month_str = evt.event.begin.datetime.strftime("%B")
            begin_year = evt.event.begin.datetime.year
            end_month_str = evt.event.end.datetime.strftime("%B")
            end_year = evt.event.end.datetime.year

            if begin_year not in monthly_events:
                monthly_events[begin_year] = {}

            if begin_month_str not in monthly_events[begin_year]:
                monthly_events[begin_year][begin_month_str] = []
            monthly_events[begin_year][begin_month_str].append(evt)

            if begin_month_str != end_month_str:
                if end_year not in monthly_events:
                    monthly_events[end_year] = {}
                if end_month_str not in monthly_events[end_year]:
                    monthly_events[end_year][end_month_str] = []
                monthly_events[end_year][end_month_str].append(evt)

        # For each month, make a plot of day vs UTC
        for year in sorted(list(monthly_events.keys())):
            for month_i in range(12):
                month = calendar.month_name[month_i + 1]
                if month not in monthly_events[year]:
                    continue

                log.info(f"Plotting {month}...")

                # Find out how many days are in the specific month
                month_number = list(calendar.month_name).index(month)
                _, days_in_month = calendar.monthrange(year, month_number)
            
                fig, axs = plt.subplots(
                    nrows=days_in_month,
                    ncols=1,
                    figsize=(20, 0.4*days_in_month),
                )
                fig.subplots_adjust(hspace=0, top=0.95)

                # fig.suptitle(f"{month} {year}")
                st = fig.suptitle(f"{month} {year}")

                # fig.autofmt_xdate()
            
                time_min = Time(f"{year}-{month_number:02d}-01T00:00:00")
                for i, ax in enumerate(axs):
                    # Plot time limits
                    ax_time_min = time_min + i * TimeDelta(1, format="jd")
                    ax_time_max = time_min + (i + 1) * TimeDelta(1, format="jd")
                    ax.set_xlim(
                        left=ax_time_min.datetime,
                        right=ax_time_max.datetime
                    )

                    hours = mdates.HourLocator(interval = 1)
                    h_fmt = mdates.DateFormatter("%H:%M")
                    ax.xaxis.set_major_locator(hours)
                    ax.xaxis.set_major_formatter(h_fmt)
                    ax.grid(axis="x", color="0.9")
                    ax.set_yticklabels([])
                    ax.set_ylabel(f"{i + 1:02d}")
                    if i != days_in_month - 1:
                       ax.set_xticklabels([])
                    # elif i == 0:
                    #     ax.set_title(f"{month} {year}")

                    for evt in monthly_events[year][month]:

                        if not self._time_intersection(
                            time1=(evt.event.begin, evt.event.end),
                            time2=(ax_time_min.datetime.replace(tzinfo=utc), ax_time_max.datetime.replace(tzinfo=utc))
                        ):
                            # Event outside of plot range
                            continue

                        # self._plot_event(ax=ax, event=event, color=kp_color)
                        evt.draw(ax=ax, fontsize=6, text_rotation=0, text_max_size=16)
                
                pdf_document.savefig(fig, bbox_inches="tight", bbox_extra_artists=[st], dpi=100)
                plt.close()

                # break

        pdf_document.close()


    # def plot(self, tmin: Time = None, tmax: Time = None, **kwargs):

    #     utc_frame = pytz.UTC

    #     time_min = Time(tmin.isot.split("T")[0])
    #     time_max = tmax
        
    #     n_subplots = int(np.ceil((tmax - tmin) / TimeDelta(1, format="jd")))

    #     fig, axs = plt.subplots(
    #         nrows=n_subplots,
    #         ncols=1,
    #         figsize=kwargs.get("figsize", (8, 3*n_subplots))
    #     )

    #     drawn_events = []
    #     def on_click(event):
    #         # Figure out the one
    #         nonlocal drawn_events
    #         for evt in drawn_events:
    #             if evt.contains(event):
    #                 print("yeah", evt.event.name)
    #     #     print('%s click: button=%d, x=%d, y=%d, xdata=%f, ydata=%f' %
    #     #         ('double' if event.dblclick else 'single', event.button,
    #     #         event.x, event.y, event.xdata, event.ydata))
        
    #     # selected_artist = None
    #     # previous_color = None
    #     # def on_key_press(event):
    #     #     nonlocal selected_artist
    #     #     if selected_artist is None:
    #     #         return
    #     #     elif event.key == "up":
    #     #         print("go to past")
    #     #         # self._shift_event_one_day_earlier(event)
    #     #     elif event.key == "down":
    #     #         print("go to future")
    #     #         # self._shift_event_one_day_later(event)
    #     #     else:
    #     #         return
    #     # def unselect_block_event():
    #     #     nonlocal selected_artist
    #     #     nonlocal previous_color
    #     #     if selected_artist is None: return
    #     #     selected_artist.set(facecolor=previous_color)
    #     #     selected_artist = None
    #     # def select_block_event(artist):
    #     #     nonlocal selected_artist
    #     #     nonlocal previous_color
    #     #     selected_artist = artist
    #     #     previous_color = selected_artist.get_facecolor()
    #     #     print(selected_artist.__dir__())
    #     #     # find the corresponding event
    #     #     selected_artist.set(facecolor="tab:red")

    #     # def onpick(event):
    #     #     if event.mouseevent.button != 1: return
    #     #     print(event)
    #     #     event_block = event.artist
    #     #     unselect_block_event()
    #     #     select_block_event(event_block)
    #     #     fig.canvas.draw()#plt.draw() #redraw

    #     # cid = fig.canvas.mpl_connect("button_press_event", on_click)
    #     # fig.canvas.mpl_connect("pick_event", onpick)
    #     # fig.canvas.mpl_connect("key_press_event", on_key_press)

    #     for i, ax in enumerate(axs):
    #         # Plot time limits
    #         ax_time_min = time_min + i * TimeDelta(1, format="jd")
    #         ax_time_max = time_min + (i + 1) * TimeDelta(1, format="jd")
    #         ax.set_xlim(
    #             left=ax_time_min.datetime,
    #             right=ax_time_max.datetime
    #         )

    #         # for kp, kp_color in zip(self.kp_names, self.kp_colors):
    
    #             #for event in self.calendars[kp].events:
    #         for evt in self.events:

    #             if not self._time_intersection(
    #                 time1=(evt.event.begin, evt.event.end),
    #                 # time2=(ax_time_min.datetime.replace(tzinfo=utc_frame), ax_time_max.datetime.replace(tzinfo=utc_frame))
    #                 time2=(time_min.datetime.replace(tzinfo=utc_frame), time_max.datetime.replace(tzinfo=utc_frame))
    #             ):
    #                 # Event outside of plot range
    #                 continue

    #             # self._plot_event(ax=ax, event=event, color=kp_color)
    #             evt.draw(ax=ax)
    #             evt.connect_to_plot_events()
    #             #drawn_events.append(evt)
        
    #     plt.tight_layout()
    #     plt.show()
    #     # plt.close("all")
    
    @staticmethod
    def _time_intersection(time1: Tuple[datetime.datetime, datetime.datetime], time2: Tuple[datetime.datetime, datetime.datetime]) -> bool:
        """Compares two time intervals (each made of a length-2 tuple)
        and returns a boolean if the two intervals intersect with each other.
        The test is performed by comparing the total duration of the time intervals
        with the duration of the interval made of the earliest start and the latest stop.
        If the total duration is shorted than the latter, it means that the intervals
        are overlapping.

        Parameters
        ----------
        time1 : Tuple[datetime.datetime, datetime.datetime]
            (start, stop)
        time2 : Tuple[datetime.datetime, datetime.datetime]
            (start, stop)

        Returns
        -------
        `bool`
            Whether the time intervals ovrlap with each other.
        """
        duration1 = time1[1] - time1[0]
        duration2 = time2[1] - time2[0]
        time_min = min(time1[0], time2[0])
        time_max = max(time1[1], time2[1])
        greatest_duration = time_max - time_min
        if duration1 + duration2 > greatest_duration:
            return True # time interval instersect each other
        else:
            return False

    def _n_intersections_event(self, event: Event) -> Tuple[int, int]:

        n_before = 0
        n_after = 0
        event_crossed = False

        for kp in self.kp_names:
            for other_event in self.calendars[kp].events:
            #     if (other_event.begin >= event.begin) & (other_event.end <= event.end):
            #         # other embedded within event
            #         n += 1
            #     elif (event.begin >= other_event.begin) & (event.end <= other_event.end):
            #         # event embedded within other
            #         n += 1
            #     elif (event.begin >= other_event.begin) & (event.begin <= other_event.end):
            #         # other overlaping start of event
            #         n += 1
            #     elif (event.end >= other_event.begin) & (event.end <= other_event.end):
            #         # other overlaping end of event
            #         n += 1
                if other_event == event:
                    event_crossed = True
                elif self._time_intersection(
                    time1=(event.begin, event.end),
                    time2=(other_event.begin, other_event.end)
                ):
                    if event_crossed:
                        n_after += 1
                    else:
                        n_before += 1
        return n_before, n_after

    # def _plot_event(self, ax: mpl.axes.Axes, event: Event, color) -> None:
    #     intersections_before, intersections_after = self._n_intersections_event(event)
    #     n_intersections = intersections_before + intersections_after + 1
        
    #     y_width = 1 / n_intersections
    #     ymin = intersections_before * y_width
    #     ymax = ymin + y_width

    #     # Show the block rectangle
    #     ax.axvspan(
    #         event.begin,
    #         event.end,
    #         ymin=ymin,
    #         ymax=ymax,
    #         facecolor=color,
    #         edgecolor="black",
    #         alpha=0.5,
    #         picker=True
    #     )

    #     # Show the observation block title
    #     x_min, x_max = ax.get_xlim()
    #     text_pos = (event.begin + (event.end - event.begin)/2)
    #     text_pos_mdate = mdates.date2num(text_pos.datetime)
    #     if (x_min <= text_pos_mdate) & (text_pos_mdate < x_max):
    #         ax.text(
    #             x=text_pos.datetime,
    #             y=ymin + y_width / 2,
    #             s=event.name + f"\n{intersections_before} -- {intersections_after}",
    #             horizontalalignment="center",
    #             verticalalignment="center",
    #             rotation=90,
    #             color="black",
    #             fontsize=8
    #         )

# ============================================================= #
# ---------------------- workbook_to_ics ---------------------- #
# def workbook_to_ics()

# ============================================================= #
# ----------------------- excel_to_ics ------------------------ #
def excel_to_ics(filename: str, output: str = None) -> None:

    # Checking input / output
    excel_extension = ".xlsx"
    calendar_extension = ".ics"
    if not filename.endswith(excel_extension):
       raise ValueError(f"{filename} does not end with '{excel_extension}'.")
    if output is None:
        output = filename.replace(excel_extension, calendar_extension)
    elif not output.endswith(calendar_extension):
        raise ValueError(f"{output} does not end with '{calendar_extension}'.")
    else:
        pass

    # Making the calendar
    calendar = Calendar()
 
    # tz = 'Europe/Paris'
    # first_day = arrow.get("2022-02-14").replace(tzinfo=tz)
    # last_day = arrow.get("2022-02-18").replace(tzinfo=tz)
    
    # for day in arrow.Arrow.range('day', first_day, last_day):
    #     event = Event()
    #     event.name = "Working on the task"
    #     event.begin = day.replace(hour=8).to('utc').datetime
    #     event.end = day.replace(hour=10).to('utc').datetime
    #     event.transparent = False
    #     calendar.events.add(event)
    
    #     event = Event()
    #     event.name = "Continue on the task?"
    #     event.begin = day.replace(hour=10).to('utc').datetime
    #     event.end = day.replace(hour=11).to('utc').datetime
    #     event.transparent = True
    #     calendar.events.add(event)

    #     with open(output, "w") as wf:
    #         wf.writelines(calendar.serialize_iter())